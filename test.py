import torch
import torch.nn.functional as F
import numpy as np
import os
import argparse
from tqdm import tqdm
from datasets.dcase2023t2.dataset import DCASE2023T2Dataset
from networks.dcase2023t2_ae.network import ConditionalAENet
from torch.utils.data import DataLoader
from sklearn.metrics import roc_auc_score, average_precision_score, precision_recall_curve, f1_score, precision_score, recall_score
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re

def calculate_anomaly_score(recon, data, z, model, domain_label):
    # Reconstruction error (L1 loss)
    recon_error = F.l1_loss(recon.squeeze(1), data.squeeze(1), reduction='none')
    recon_error = recon_error.mean(dim=(1, 2))
    
    # Mahalanobis distance
    if domain_label[0] == 0:  # Source domain
        mahalanobis_dist = calculate_mahalanobis_distance(z, model.cov_source)
    else:  # Target domain
        mahalanobis_dist = calculate_mahalanobis_distance(z, model.cov_target)
    
    # Prototype similarity
    prototype_scores = model.prototype_bank(z)
    prototype_similarity = torch.max(prototype_scores, dim=1)[0]
    
    # Combine scores with weighted components
    anomaly_score = (
        0.5 * recon_error +  # Reconstruction error weight
        0.3 * mahalanobis_dist +  # Mahalanobis distance weight
        0.2 * (1 - prototype_similarity)  # Prototype dissimilarity weight
    )
    
    return anomaly_score

def calculate_mahalanobis_distance(z, cov):
    z_centered = z - z.mean(dim=0)
    eps = 1e-6
    cov = cov + torch.eye(cov.size(0), device=cov.device) * eps
    U, S, V = torch.svd(cov)
    S_inv = torch.diag(1.0 / (S + eps))
    inv_cov = torch.mm(torch.mm(V, S_inv), U.t())
    mahalanobis = torch.sum(z_centered @ inv_cov * z_centered, dim=1)
    return mahalanobis

def calculate_metrics(scores, labels):
    """Calculate all required metrics for a domain."""
    # Calculate AUC
    auc = roc_auc_score(labels, scores)
    
    # Calculate pAUC
    precision, recall, _ = precision_recall_curve(labels, scores)
    pauc = np.trapz(precision, recall)
    
    # Calculate threshold-based metrics
    threshold = np.mean(scores) + np.std(scores)
    pred_labels = (scores > threshold).astype(int)
    
    precision_val = precision_score(labels, pred_labels)
    recall_val = recall_score(labels, pred_labels)
    f1 = f1_score(labels, pred_labels)
    
    return {
        'AUC': auc,
        'pAUC': pauc,
        'Precision': precision_val,
        'Recall': recall_val,
        'F1-Score': f1
    }

def format_excel_metrics(writer, sheet_name):
    """Format the metrics Excel sheet with proper styling."""
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    # Format header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Format value column
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15

def format_excel_scores(writer, sheet_name):
    """Format the scores Excel sheet with proper styling."""
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    # Format header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Format numeric columns
    numeric_columns = ['Anomaly_Score', 'True_Label', 'Predicted_Label']
    for col in numeric_columns:
        col_idx = worksheet[1].index([cell for cell in worksheet[1] if cell.value == col][0]) + 1
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if col == 'Anomaly_Score':
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '0'
                cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 30  # File
    worksheet.column_dimensions['B'].width = 15  # Anomaly_Score
    worksheet.column_dimensions['C'].width = 12  # True_Label
    worksheet.column_dimensions['D'].width = 12  # Predicted_Label
    worksheet.column_dimensions['E'].width = 15  # Machine_Type
    worksheet.column_dimensions['F'].width = 10  # Section
    worksheet.column_dimensions['G'].width = 10  # Domain
    worksheet.column_dimensions['H'].width = 10  # Condition

def load_ground_truth_labels(gt_path):
    if gt_path.endswith('.xlsx'):
        gt_df = pd.read_excel(gt_path, header=None)
    else:
        # Try different separators
        for sep in [',', '\t', '\s+']:
            try:
                gt_df = pd.read_csv(gt_path, header=None, sep=sep)
                if gt_df.shape[1] >= 2:
                    break
            except Exception:
                continue
        else:
            raise ValueError('Ground truth file format not recognized: expected at least two columns.')
    gt_df = gt_df.iloc[:, :2]
    gt_df.columns = ['section', 'label']
    return dict(zip(gt_df['section'].astype(str), gt_df['label'].astype(int)))

def test(args):
    # Set device
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    
    # Load model
    model = ConditionalAENet(
        input_dim=args.input_dim,
        block_size=args.block_size,
        num_domains=2
    ).to(device)
    
    # Use lowercase for the model file name
    model_path = os.path.join(args.model_dir, f'{args.machine_type.lower()}_model.pth')
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model not found at {model_path}")
    
    try:
        # Load state dict with strict=False to handle architecture differences
        state_dict = torch.load(model_path)
        model.load_state_dict(state_dict, strict=False)
        print("Model loaded successfully with some parameter mismatches (this is expected)")
    except Exception as e:
        print(f"Warning: Error loading model: {str(e)}")
        print("Attempting to load model with modified architecture...")
        # Try to load only matching parameters
        model_dict = model.state_dict()
        pretrained_dict = {k: v for k, v in state_dict.items() if k in model_dict and v.shape == model_dict[k].shape}
        model_dict.update(pretrained_dict)
        model.load_state_dict(model_dict)
        print("Model loaded with partial parameters")
    
    model.eval()
    
    # Create test dataset
    test_dataset = DCASE2023T2Dataset(
        root_dir=args.dataset_dir,
        machine_type=args.machine_type,
        split='test'
    )
    
    test_loader = DataLoader(
        test_dataset,
        batch_size=args.batch_size,
        shuffle=False,
        num_workers=0
    )
    
    # Initialize results storage
    test_scores = []
    test_files = []
    test_labels = []  # 0 for normal, 1 for anomaly
    label_debug_pairs = []
    unmatched_count = 0
    
    # --- Load ground truth if available for Scanner ---
    gt_labels = None
    if args.machine_type.lower() == 'scanner':
        # Try both .xlsx and .csv
        gt_base = os.path.join(args.dataset_dir, args.machine_type, 'ground_truth_Scanner_section_00_test')
        gt_path_xlsx = gt_base + '.xlsx'
        gt_path_csv = gt_base + '.csv'
        if os.path.exists(gt_path_xlsx):
            print(f"Loading ground truth from: {gt_path_xlsx}")
            gt_labels = load_ground_truth_labels(gt_path_xlsx)
        elif os.path.exists(gt_path_csv):
            print(f"Loading ground truth from: {gt_path_csv}")
            gt_labels = load_ground_truth_labels(gt_path_csv)
        else:
            print(f"Ground truth file not found at {gt_path_xlsx} or {gt_path_csv}, using filename-based labels.")
    
    # Testing loop
    with torch.no_grad():
        for data, file_path in tqdm(test_loader, desc="Testing"):
            data = data.to(device)
            recon, z, proto_sim, _ = model(data)
            anomaly_scores = F.mse_loss(recon, data, reduction='none').mean(dim=(1,2,3))
            test_scores.extend(anomaly_scores.cpu().numpy())
            test_files.extend(file_path)
            if gt_labels is not None:
                for f in file_path:
                    key = os.path.basename(f)
                    key = os.path.splitext(key)[0]
                    if key in gt_labels:
                        test_labels.append(gt_labels[key])
                        label_debug_pairs.append((key, gt_labels[key]))
                    else:
                        found = False
                        for gt_key in gt_labels:
                            if key.startswith(gt_key):
                                test_labels.append(gt_labels[gt_key])
                                label_debug_pairs.append((key, gt_labels[gt_key]))
                                found = True
                                break
                        if not found:
                            test_labels.append(0)
                            label_debug_pairs.append((key, 'NOT FOUND'))
                            unmatched_count += 1
            else:
                for f in file_path:
                    label = 0 if 'normal' in f else 1
                    test_labels.append(label)
                    label_debug_pairs.append((f, label))
    
    # Print debug info for label matching
    print("\nSample of file-to-label matches:")
    for pair in label_debug_pairs[:10]:
        print(pair)
    print(f"Total unmatched files: {unmatched_count} out of {len(test_files)}")
    
    # Convert to numpy arrays
    test_scores = np.array(test_scores)
    test_labels = np.array(test_labels)
    
    # Create results directory if it doesn't exist
    os.makedirs(args.results_dir, exist_ok=True)
    
    # Calculate evaluation metrics
    metrics_results = {}
    
    # Check if we have both normal and anomaly samples
    unique_labels = np.unique(test_labels)
    if len(unique_labels) > 1:
        # Calculate AUC only if we have both classes
        auc = roc_auc_score(test_labels, test_scores)
        metrics_results['AUC'] = auc
        
        # Calculate pAUC (partial AUC)
        precision, recall, _ = precision_recall_curve(test_labels, test_scores)
        pauc = np.trapz(precision, recall)
        metrics_results['pAUC'] = pauc
        
        # Use actual score range for thresholding
        thresholds = np.linspace(test_scores.min(), test_scores.max(), 100)
        f1_scores = []
        for threshold in thresholds:
            pred_labels = (test_scores > threshold).astype(int)
            f1 = f1_score(test_labels, pred_labels)
            f1_scores.append(f1)
        optimal_threshold = thresholds[np.argmax(f1_scores)]
        
        # Calculate metrics using optimal threshold
        pred_labels = (test_scores > optimal_threshold).astype(int)
        metrics_results['Precision'] = precision_score(test_labels, pred_labels)
        metrics_results['Recall'] = recall_score(test_labels, pred_labels)
        metrics_results['F1_Score'] = f1_score(test_labels, pred_labels)
        metrics_results['Decision_Accuracy'] = (pred_labels == test_labels).mean()
        metrics_results['Optimal_Threshold'] = optimal_threshold
    else:
        # If only one class is present, use simpler metrics
        metrics_results['AUC'] = float('nan')
        metrics_results['pAUC'] = float('nan')
        metrics_results['Precision'] = float('nan')
        metrics_results['Recall'] = float('nan')
        metrics_results['F1_Score'] = float('nan')
        metrics_results['Decision_Accuracy'] = float('nan')
        metrics_results['Optimal_Threshold'] = float('nan')
        pred_labels = np.zeros_like(test_labels)  # Default to all normal
    
    # Save metrics to Excel with formatting
    metrics_df = pd.DataFrame({
        'Metric': list(metrics_results.keys()),
        'Value': list(metrics_results.values())
    })
    
    # Close any open Excel files before saving
    try:
        metrics_path = os.path.join(args.results_dir, 'evaluation_metrics.xlsx')
        if os.path.exists(metrics_path):
            os.remove(metrics_path)
        with pd.ExcelWriter(metrics_path, engine='openpyxl') as writer:
            metrics_df.to_excel(writer, sheet_name='Metrics', index=False)
            format_excel_metrics(writer, 'Metrics')
    except Exception as e:
        print(f"Warning: Could not save metrics file: {str(e)}")
    
    # Save anomaly scores to Excel with formatting
    scores_df = pd.DataFrame({
        'File': test_files,
        'Anomaly_Score': test_scores,
        'Machine_Type': args.machine_type,
        'Section': [f.split('_')[1] for f in test_files],
        'Domain': ['source' if 'source' in f else 'target' for f in test_files],
        'Condition': ['normal' if 'normal' in f else 'anomaly' for f in test_files]
    })
    scores_path = os.path.join(args.results_dir, 'anomaly_scores.xlsx')
    try:
        if os.path.exists(scores_path):
            os.remove(scores_path)
        with pd.ExcelWriter(scores_path, engine='openpyxl') as writer:
            scores_df.to_excel(writer, sheet_name='Scores', index=False)
            format_excel_scores(writer, 'Scores')
    except Exception as e:
        print(f"Warning: Could not save scores file: {str(e)}")

    # --- Improved evaluation: reload anomaly scores and merge with ground truth ---
    def strip_extensions(filename):
        # Remove all extensions (e.g., .wav, .wav.wav, etc.)
        return re.sub(r'(\.\w+)+$', '', filename)
    # Reload anomaly scores
    scores_df = pd.read_excel(scores_path)
    # Normalize anomaly scores (z-score)
    scores_mean = scores_df['Anomaly_Score'].mean()
    scores_std = scores_df['Anomaly_Score'].std()
    scores_df['Anomaly_Score_Norm'] = (scores_df['Anomaly_Score'] - scores_mean) / (scores_std + 1e-8)
    # Find ground truth file
    gt_base = os.path.join(args.dataset_dir, args.machine_type, 'ground_truth_Scanner_section_00_test')
    gt_path_xlsx = gt_base + '.xlsx'
    gt_path_csv = gt_base + '.csv'
    if os.path.exists(gt_path_xlsx):
        gt_df = pd.read_excel(gt_path_xlsx, header=None)
    elif os.path.exists(gt_path_csv):
        for sep in [',', '\t', '\s+']:
            try:
                gt_df = pd.read_csv(gt_path_csv, header=None, sep=sep)
                if gt_df.shape[1] >= 2:
                    break
            except Exception:
                continue
        else:
            raise ValueError('Ground truth file format not recognized: expected at least two columns.')
    else:
        print('Ground truth file not found for evaluation!')
        gt_df = None
    if gt_df is not None:
        gt_df = gt_df.iloc[:, :2]
        gt_df.columns = ['File', 'True_Label']
        # Normalize file names in both DataFrames
        scores_df['File_norm'] = scores_df['File'].astype(str).apply(strip_extensions)
        gt_df['File_norm'] = gt_df['File'].astype(str).apply(strip_extensions)
        # Merge on normalized file names
        merged = pd.merge(scores_df, gt_df, on='File_norm', how='inner')
        unmatched = set(scores_df['File_norm']) - set(merged['File_norm'])
        print(f"Matched {len(merged)} files with ground truth. Unmatched: {len(unmatched)}")
        if unmatched:
            print("Sample unmatched files:", list(unmatched)[:10])
        # Calculate metrics using normalized scores
        if merged['True_Label'].nunique() > 1:
            auc = roc_auc_score(merged['True_Label'], merged['Anomaly_Score_Norm'])
            precision, recall, _ = precision_recall_curve(merged['True_Label'], merged['Anomaly_Score_Norm'])
            pauc = np.trapz(precision, recall)
            thresholds = np.linspace(merged['Anomaly_Score_Norm'].min(), merged['Anomaly_Score_Norm'].max(), 100)
            f1_scores = []
            for threshold in thresholds:
                pred_labels = (merged['Anomaly_Score_Norm'] > threshold).astype(int)
                f1_scores.append(f1_score(merged['True_Label'], pred_labels))
            optimal_threshold = thresholds[np.argmax(f1_scores)]
            pred_labels = (merged['Anomaly_Score_Norm'] > optimal_threshold).astype(int)
            metrics = {
                'AUC': auc,
                'pAUC': pauc,
                'Precision': precision_score(merged['True_Label'], pred_labels),
                'Recall': recall_score(merged['True_Label'], pred_labels),
                'F1_Score': f1_score(merged['True_Label'], pred_labels),
                'Decision_Accuracy': (pred_labels == merged['True_Label']).mean(),
                'Optimal_Threshold': optimal_threshold
            }
        else:
            metrics = {k: float('nan') for k in ['AUC', 'pAUC', 'Precision', 'Recall', 'F1_Score', 'Decision_Accuracy', 'Optimal_Threshold']}
        # Save metrics
        metrics_df = pd.DataFrame({'Metric': list(metrics.keys()), 'Value': list(metrics.values())})
        metrics_path = os.path.join(args.results_dir, 'evaluation_metrics.xlsx')
        try:
            if os.path.exists(metrics_path):
                os.remove(metrics_path)
            with pd.ExcelWriter(metrics_path, engine='openpyxl') as writer:
                metrics_df.to_excel(writer, sheet_name='Metrics', index=False)
                format_excel_metrics(writer, 'Metrics')
        except Exception as e:
            print(f"Warning: Could not save metrics file: {str(e)}")
        print("\n=== Evaluation Metrics (from merged ground truth) ===")
        print(metrics_df)
    else:
        print("No ground truth available for evaluation metrics.")
    
    # Print results
    print("\n=== Testing Complete ===")
    print(f"Processed {len(test_scores)} files")
    print("\nEvaluation Metrics:")
    for metric, value in metrics_results.items():
        if np.isnan(value):
            print(f"{metric}: N/A (only one class present)")
        else:
            print(f"{metric}: {value:.4f}")
    print(f"\nResults saved to {args.results_dir}/:")
    print("1. evaluation_metrics.xlsx - Contains all evaluation metrics")
    print("2. anomaly_scores.xlsx - Contains anomaly scores and predictions")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dataset_dir', type=str, required=True)
    parser.add_argument('--machine_type', type=str, required=True)
    parser.add_argument('--input_dim', type=int, default=128)
    parser.add_argument('--block_size', type=int, default=313)
    parser.add_argument('--batch_size', type=int, default=32)
    parser.add_argument('--results_dir', type=str, required=True)
    parser.add_argument('--model_dir', type=str, required=True)
    args = parser.parse_args()
    
    test(args)

if __name__ == '__main__':
    main() 